#!/usr/bin/env python3
"""
One-time import for the influencer tracker.

Reads the CK Influencer Budget Tracker spreadsheet and upserts:
  - influencer_products  (style_code, product_name, brand, cost_price, rrp, cost_ratio)
  - influencer_budgets   (brand, month_key, budget)  for FY26/27

Cost prices are written to Supabase only (server-side). Run AFTER creating the
tables (see the SQL in the Influencer tab's setup card).

Usage:  python3 scripts/import_influencer.py "/path/to/CK - Influencer Budget Tracker.xlsx"
Env:    NEXT_PUBLIC_SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY, NEXT_PUBLIC_SUPABASE_ANON_KEY
"""
import sys, os, json, ssl, urllib.request, urllib.error

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ENV = os.path.join(BASE, ".env.local")

def load_env():
    if not os.path.exists(ENV): return
    for line in open(ENV):
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line: continue
        k, _, v = line.partition("="); os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

load_env()
URL = os.environ.get("NEXT_PUBLIC_SUPABASE_URL", "")
SVC = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
ANON = os.environ.get("NEXT_PUBLIC_SUPABASE_ANON_KEY", SVC)

# Normalise the messy brand strings to the canonical budget brands
BRAND_FIX = {
    "wonderfold": "WonderFold", "gaia baby": "Gaia", "gaia": "Gaia", "zazu": "Zazu",
    "matchstick monkey": "Matchstick Monkey", "miamily": "MiaMily", "mamave": "Mamave",
    "uppababy": "UPPAbaby", "frida": "Frida", "nanit": "Nanit", "hannie": "Hannie",
    "magic": "Magic", "smartrike": "SmarTrike",
}
def norm_brand(b):
    if b is None: return None
    s = str(b).strip()
    if s in ("", "0"): return None
    return BRAND_FIX.get(s.lower(), s)

MONTH_KEYS = [f"2026-{m:02d}" for m in range(7, 13)] + [f"2027-{m:02d}" for m in range(1, 7)]

def upsert(table, rows, on_conflict):
    if not rows: return
    url = f"{URL}/rest/v1/{table}?on_conflict={on_conflict}"
    req = urllib.request.Request(url, data=json.dumps(rows).encode(), method="POST")
    req.add_header("Content-Type", "application/json")
    req.add_header("Authorization", f"Bearer {SVC}")
    req.add_header("apikey", ANON)
    req.add_header("Prefer", "resolution=merge-duplicates,return=minimal")
    try:
        with urllib.request.urlopen(req, context=ssl.create_default_context(), timeout=60) as r:
            return r.status
    except urllib.error.HTTPError as e:
        print(f"  ✗ {table}: {e.code} {e.read().decode()[:300]}"); sys.exit(1)

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 scripts/import_influencer.py <xlsx path>"); sys.exit(1)
    path = sys.argv[1]
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    # ── Products from CONFIG ──
    cfg = wb["CONFIG"]
    products, seen = [], set()
    for r in range(2, cfg.max_row + 1):
        name = cfg.cell(r, 1).value; code = cfg.cell(r, 2).value
        if not code or not name: continue
        code = str(code).strip()
        if code in seen: continue
        seen.add(code)
        products.append({
            "style_code": code, "product_name": str(name).strip(), "brand": norm_brand(cfg.cell(r, 3).value),
            "cost_price": cfg.cell(r, 4).value, "rrp": cfg.cell(r, 5).value, "cost_ratio": cfg.cell(r, 6).value,
        })
    print(f"Products: {len(products)}")
    for i in range(0, len(products), 200):
        upsert("influencer_products", products[i:i+200], "style_code")

    # ── Budgets from Overall Budget (Brand × month grid) ──
    ob = wb["Overall Budget"]
    # header row 4: B..M = Jul 2026 .. Jun 2027
    budgets = []
    for r in range(5, ob.max_row + 1):
        brand = norm_brand(ob.cell(r, 1).value)
        if not brand: continue
        for ci, mk in enumerate(MONTH_KEYS):
            v = ob.cell(r, 2 + ci).value
            if v not in (None, ""):
                budgets.append({"brand": brand, "month_key": mk, "budget": float(v)})
    print(f"Budgets: {len(budgets)} brand-months")
    for i in range(0, len(budgets), 200):
        upsert("influencer_budgets", budgets[i:i+200], "brand,month_key")

    print("✅ Import complete")

if __name__ == "__main__":
    main()
