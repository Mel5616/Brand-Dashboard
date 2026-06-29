#!/usr/bin/env python3
"""Parse the Monthly Promo Tracker (xlsx) and upsert sale periods into `promotions`.

Usage: python3 scripts/load_promotions.py "/path/to/Promo Calendar.xlsx"

Reads the "Monthly Promo Tracker" sheet: each quarter block has a "Brand" header
row whose columns hold week-start dates, then one row per brand. A non-empty cell
means that brand is on sale that week; the cell text is the retailer/channel.
Consecutive on-sale weeks with the same channel merge into one period.

Pricing/notes entered in the dashboard are NEVER overwritten — the upsert only
touches brand/date/channel columns. Stale sheet-sourced rows are pruned.
"""
import os, sys, json, datetime, urllib.request, urllib.parse
import openpyxl

SB_URL = os.environ["NEXT_PUBLIC_SUPABASE_URL"].rstrip("/")
SB_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]

# sheet brand name -> brands.id
BRANDS = {0:"nanit",1:"magic",2:"hannie",3:"gaiababy",4:"wonderfold",5:"uppababy",
          6:"zazu",7:"miamily",8:"frida",10:"matchstickmonkey",11:"mamave"}
def norm(s): return "".join(str(s).lower().split())
def brand_id_for(name):
    n = norm(name)
    for bid, bn in BRANDS.items():
        if n == bn or bn.startswith(n) or n.startswith(bn):
            return bid
    return None

# Cell shading -> tier: theme7 (blue #0F9ED5) = Tier 1, theme9 (green #4EA72E) = Tier 2.
def tier_of(cell):
    f = cell.fill
    if not (f and f.patternType): return None
    t = getattr(f.fgColor, "theme", None)
    return 1 if t == 7 else 2 if t == 9 else None

def parse(path):
    wv = openpyxl.load_workbook(path, data_only=True)
    ws = wv["Monthly Promo Tracker"]
    st = openpyxl.load_workbook(path)["Monthly Promo Tracker"]  # for fill colours
    rows = []
    headers = [r for r in range(1, ws.max_row+1) if str(ws.cell(r,1).value).strip().lower() == "brand"]
    for hr in headers:
        cols = []
        for c in range(2, ws.max_column+1):
            v = ws.cell(hr, c).value
            if isinstance(v, (datetime.datetime, datetime.date)):
                cols.append((c, v.date() if isinstance(v, datetime.datetime) else v))
        r = hr + 1
        while r <= ws.max_row:
            b = ws.cell(r, 1).value
            if b is None or str(b).strip() == "" or str(b).strip().lower() == "brand":
                break
            brand = str(b).strip()
            for (c, wkstart) in cols:
                t = ws.cell(r, c).value
                if t is not None and str(t).strip() != "":
                    chan = " ".join(str(t).split())
                    rows.append([brand, wkstart, wkstart + datetime.timedelta(days=6), chan, tier_of(st.cell(r, c))])
            r += 1
    # merge contiguous weeks for same brand+channel+tier
    rows.sort(key=lambda p: (p[0], p[3], p[1]))
    merged = []
    for b, s, e, ch, ti in rows:
        if merged and merged[-1][0] == b and merged[-1][3] == ch and merged[-1][4] == ti and (s - merged[-1][2]).days <= 1:
            merged[-1][2] = e
        else:
            merged.append([b, s, e, ch, ti])
    return merged

def parse_lines(path):
    """Parse the Promo Details sheet into marketing-facing product-promo lines."""
    d = openpyxl.load_workbook(path, data_only=True)["Promo Details"]
    def num(v):
        try: return float(v) if v not in (None, "") else None
        except Exception: return None
    def dt(v): return v.date().isoformat() if isinstance(v, (datetime.datetime, datetime.date)) else None
    out = []
    for r in range(3, d.max_row + 1):
        if not d.cell(r, 2).value:  # no brand → skip
            continue
        out.append({
            "customer": d.cell(r, 1).value, "brand": str(d.cell(r, 2).value).strip(),
            "brand_id": brand_id_for(d.cell(r, 2).value), "sku": d.cell(r, 3).value,
            "promo_name": d.cell(r, 4).value, "product": d.cell(r, 5).value,
            "category": d.cell(r, 9).value, "month": d.cell(r, 10).value,
            "tier": int(d.cell(r, 11).value) if str(d.cell(r, 11).value).strip() in ("1", "2") else None,
            "start_date": dt(d.cell(r, 12).value), "end_date": dt(d.cell(r, 13).value),
            "days": int(num(d.cell(r, 14).value)) if num(d.cell(r, 14).value) else None,
            "rrp": num(d.cell(r, 16).value), "current_price": num(d.cell(r, 17).value),
            "promo_price": num(d.cell(r, 19).value), "discount_rrp": num(d.cell(r, 20).value),
            "source": "sheet",
        })
    return out

def req(method, path, body=None, prefer=None):
    h = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}", "Content-Type": "application/json"}
    if prefer: h["Prefer"] = prefer
    data = json.dumps(body).encode() if body is not None else None
    r = urllib.request.Request(f"{SB_URL}/rest/v1/{path}", data=data, headers=h, method=method)
    with urllib.request.urlopen(r) as resp:
        return resp.status, resp.read().decode()

def load_from_path(path):
    CUTOFF = "2026-07-01"  # FY26 — drop any promo ending before this
    periods = parse(path)
    payload = []
    for brand, s, e, ch, ti in periods:
        if e.isoformat() < CUTOFF:
            continue
        payload.append({
            "brand_id": brand_id_for(brand), "brand": brand,
            "period_start": s.isoformat(), "period_end": e.isoformat(),
            "channel": ch, "tier": ti, "source": "sheet", "updated_at": datetime.datetime.utcnow().isoformat(),
        })
    # upsert calendar (price/note untouched), then prune stale sheet rows
    st, _ = req("POST", "promotions?on_conflict=brand,period_start,channel", payload,
                "resolution=merge-duplicates,return=minimal")
    keep = {(p["brand"], p["period_start"], p["channel"]) for p in payload}
    _, existing = req("GET", "promotions?select=id,brand,period_start,channel&source=eq.sheet")
    stale = [row["id"] for row in json.loads(existing) if (row["brand"], row["period_start"], row["channel"]) not in keep]
    for i in stale:
        req("DELETE", f"promotions?id=eq.{i}", prefer="return=minimal")
    print(f"calendar: upserted {len(payload)} promo periods (HTTP {st}); pruned {len(stale)} stale")

    # promo_lines have no manual edits → replace all sheet rows (FY26+ only)
    lines = [l for l in parse_lines(path) if l.get("end_date") and l["end_date"] >= CUTOFF]
    req("DELETE", "promo_lines?source=eq.sheet", prefer="return=minimal")
    for i in range(0, len(lines), 200):
        req("POST", "promo_lines", lines[i:i+200], "return=minimal")
    print(f"lines: loaded {len(lines)} product-promo lines (FY26+)")

    # D2C plan — dedupe lines to one row per brand+product+window; keep status/note.
    groups = {}
    for l in lines:
        key = (l["brand"], l.get("sku") or l.get("product") or "", l["start_date"], l["end_date"])
        if None in (l["start_date"], l["end_date"]):
            continue
        g = groups.get(key)
        if g is None or (l.get("promo_price") is not None and (g["promo_price"] is None or l["promo_price"] < g["promo_price"])):
            best = {"rrp": l.get("rrp"), "promo_price": l.get("promo_price"), "discount_rrp": l.get("discount_rrp"), "tier": l.get("tier")}
        else:
            best = {k: g[k] for k in ("rrp", "promo_price", "discount_rrp", "tier")}
        retailers = sorted(set((g["_ret"] if g else []) + ([l["customer"]] if l.get("customer") else [])))
        groups[key] = {
            "brand": l["brand"], "brand_id": l.get("brand_id"), "sku": key[1], "product": l.get("product"),
            "period_start": l["start_date"], "period_end": l["end_date"],
            **best, "retailers": ", ".join(retailers), "_ret": retailers,
            "source": "sheet", "updated_at": datetime.datetime.utcnow().isoformat(),
        }
    d2c = [{k: v for k, v in g.items() if k != "_ret"} for g in groups.values()]
    st3, _ = req("POST", "d2c_promos?on_conflict=brand,sku,period_start,period_end", d2c,
                 "resolution=merge-duplicates,return=minimal")
    keepd = {(g["brand"], g["sku"], g["period_start"], g["period_end"]) for g in d2c}
    _, ex = req("GET", "d2c_promos?select=id,brand,sku,period_start,period_end&source=eq.sheet")
    staled = [r["id"] for r in json.loads(ex) if (r["brand"], r["sku"], r["period_start"], r["period_end"]) not in keepd]
    for i in staled:
        req("DELETE", f"d2c_promos?id=eq.{i}", prefer="return=minimal")
    print(f"d2c: upserted {len(d2c)} D2C promos (HTTP {st3}); pruned {len(staled)} stale")

def download_via_graph():
    """Download the Promo Calendar xlsx from SharePoint via Microsoft Graph
    (app-only client-credentials). Returns a temp file path, or None if not configured."""
    import base64, tempfile
    tenant = os.environ.get("MS_TENANT_ID")
    client = os.environ.get("MS_CLIENT_ID")
    secret = os.environ.get("MS_CLIENT_SECRET")
    share_url = os.environ.get("MS_PROMO_SHARE_URL")
    if not (tenant and client and secret and share_url):
        return None
    # 1) token
    body = urllib.parse.urlencode({
        "client_id": client, "client_secret": secret,
        "scope": "https://graph.microsoft.com/.default", "grant_type": "client_credentials",
    }).encode()
    tok_req = urllib.request.Request(f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token", data=body,
                                     headers={"Content-Type": "application/x-www-form-urlencoded"}, method="POST")
    with urllib.request.urlopen(tok_req) as r:
        token = json.loads(r.read().decode())["access_token"]
    # 2) resolve share link -> driveItem content
    enc = "u!" + base64.urlsafe_b64encode(share_url.encode()).decode().rstrip("=")
    dl = urllib.request.Request(f"https://graph.microsoft.com/v1.0/shares/{enc}/driveItem/content",
                                headers={"Authorization": f"Bearer {token}"})
    with urllib.request.urlopen(dl) as r:
        data = r.read()
    fd, tmp = tempfile.mkstemp(suffix=".xlsx")
    with os.fdopen(fd, "wb") as f:
        f.write(data)
    return tmp

def main():
    # Explicit local path wins (manual run); otherwise pull from SharePoint via Graph.
    if len(sys.argv) > 1:
        load_from_path(sys.argv[1]); return
    tmp = None
    try:
        tmp = download_via_graph()
    except Exception as e:
        print(f"promo sync: Graph download failed: {e}"); return
    if not tmp:
        print("promo sync: MS_* env not set — skipping SharePoint pull."); return
    print("promo sync: downloaded Promo Calendar from SharePoint")
    load_from_path(tmp)

if __name__ == "__main__":
    main()
